import serial
import requests
import time
import os

from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

# =====================================================
# CONFIGURATION
# =====================================================

BOT_TOKEN = "YOUR_BOT_TOKEN"
CHAT_ID = "YOUR_CHAT_ID"

SERIAL_PORT = "COM7"
BAUD_RATE = 115200

EXCEL_FILE = r"C:\Users\Public\traffic_analysis.xlsx"

# =====================================================
# TELEGRAM FUNCTION
# =====================================================

def send_telegram(message):

    try:

        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"

        requests.post(
            url,
            json={
                "chat_id": CHAT_ID,
                "text": message
            },
            timeout=5
        )

        print("Telegram Sent")

    except Exception as e:

        print("Telegram Error :", e)

# =====================================================
# CREATE EXCEL FILE
# =====================================================

def create_excel():

    if not os.path.exists(EXCEL_FILE):

        wb = Workbook()

        ws = wb.active

        ws.title = "Traffic Log"

        ws.append([
            "Date",
            "Time",
            "Signal",
            "State",
            "Duration(sec)",
            "Mode"
        ])

        wb.save(EXCEL_FILE)

        print("Excel File Created")

# =====================================================
# SAVE DATA TO EXCEL
# =====================================================

def save_excel(signal,
               state,
               duration,
               mode):

    wb = load_workbook(EXCEL_FILE)

    ws = wb.active

    now = datetime.now()

    ws.append([
        now.strftime("%d-%m-%Y"),
        now.strftime("%H:%M:%S"),
        signal,
        state,
        round(duration,2),
        mode
    ])

    wb.save(EXCEL_FILE)

# =====================================================
# INITIALIZE EXCEL
# =====================================================

create_excel()

# =====================================================
# SERIAL CONNECTION
# =====================================================

ser = serial.Serial(
    SERIAL_PORT,
    BAUD_RATE,
    timeout=1
)

print("===================================")
print(" SMART TRAFFIC MONITOR STARTED ")
print("===================================")
print("Connected to :", SERIAL_PORT)
print()

# =====================================================
# GLOBAL VARIABLES
# =====================================================

current_state = None

state_start = None

priority_mode = False

priority_signal = ""

pedestrian_event = False

minute_start = datetime.now()

summary_data = []

signal1_extensions = 0

signal2_extensions = 0